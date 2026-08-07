#!/usr/bin/env python3
"""Export the requirements traceability matrix to a formatted spreadsheet.

The CSV is the source of truth. This produces a readable, filterable view of
exactly that file: nothing is recomputed, reordered, or relabelled here, so the
spreadsheet can never disagree with the validated matrix.

The status column is colour-coded (green production ready, amber implemented
but not fully verified, red blocked or not implemented, grey not applicable), a
frozen header row and an autofilter make 176 rows navigable, and long-form
columns wrap instead of running off the page.

openpyxl (MIT) is a tooling dependency in the ``dev`` extra, not a product
runtime dependency, so the zero-dependency core is unaffected.

Usage:
    python scripts/export_matrix_xlsx.py
    python scripts/export_matrix_xlsx.py --check   # verify it matches the CSV
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "docs" / "REQUIREMENTS_TRACEABILITY_MATRIX.csv"
XLSX_PATH = ROOT / "docs" / "REQUIREMENTS_TRACEABILITY_MATRIX.xlsx"

# Fill and font per status. Colours follow the conventional spreadsheet palette
# so the meaning is readable without a legend.
STATUS_STYLE = {
    "PRODUCTION READY": ("C6EFCE", "006100"),
    "IMPLEMENTED BUT NOT FULLY VERIFIED": ("FFEB9C", "9C5700"),
    "PARTIAL": ("FFEB9C", "9C5700"),
    "BLOCKED BY EXTERNAL PLATFORM": ("FFC7CE", "9C0006"),
    "NOT IMPLEMENTED": ("FFC7CE", "9C0006"),
    "NOT APPLICABLE": ("D9D9D9", "3F3F3F"),
}

# Column width in characters, and whether the column wraps.
COLUMN_LAYOUT = {
    "id": (9, False),
    "area": (26, False),
    "source_ref": (16, False),
    "requirement": (55, True),
    "implementation_files": (40, True),
    "tests": (34, True),
    "acceptance_test": (38, True),
    "status": (34, False),
    "evidence_path": (30, False),
    "licence_impact": (24, True),
    "remaining_limitation": (50, True),
}
DEFAULT_LAYOUT = (24, True)


def _read_rows() -> tuple[list[str], list[dict]]:
    with CSV_PATH.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    if not fieldnames:
        raise SystemExit(f"export-matrix: {CSV_PATH} has no header; refusing to write an empty sheet")
    if not rows:
        raise SystemExit(f"export-matrix: {CSV_PATH} has no rows; refusing to write an empty sheet")
    return fieldnames, rows


def build() -> tuple[int, dict[str, int]]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover - exercised only without the extra
        raise SystemExit(
            "export-matrix: openpyxl is required: pip install -e '.[dev]'"
        ) from e

    fieldnames, rows = _read_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Traceability"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(vertical="center", horizontal="left")

    counts: dict[str, int] = {}
    status_index = fieldnames.index("status") + 1 if "status" in fieldnames else None
    for row in rows:
        ws.append([row.get(name, "") for name in fieldnames])
        status = (row.get("status") or "").strip()
        counts[status] = counts.get(status, 0) + 1
        r = ws.max_row
        for c in range(1, len(fieldnames) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            _name = fieldnames[c - 1]
            _width, wrap = COLUMN_LAYOUT.get(_name, DEFAULT_LAYOUT)
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
        if status_index is not None:
            style = STATUS_STYLE.get(status)
            if style:
                fill_colour, font_colour = style
                cell = ws.cell(row=r, column=status_index)
                cell.fill = PatternFill("solid", fgColor=fill_colour)
                cell.font = Font(bold=True, color=font_colour)

    for i, name in enumerate(fieldnames, start=1):
        width, _wrap = COLUMN_LAYOUT.get(name, DEFAULT_LAYOUT)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{ws.max_row}"

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    return len(rows), counts


def check() -> int:
    """Confirm the spreadsheet on disk still matches the CSV row for row."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise SystemExit("export-matrix: openpyxl is required: pip install -e '.[dev]'") from e

    if not XLSX_PATH.exists():
        print(f"export-matrix: {XLSX_PATH.name} is missing", file=sys.stderr)
        return 1
    fieldnames, rows = _read_rows()
    ws = load_workbook(XLSX_PATH, read_only=True).active
    values = list(ws.values)
    if list(values[0]) != fieldnames:
        print("export-matrix: header mismatch between CSV and spreadsheet", file=sys.stderr)
        return 1
    if len(values) - 1 != len(rows):
        print(
            f"export-matrix: row count mismatch: csv {len(rows)}, spreadsheet {len(values) - 1}",
            file=sys.stderr,
        )
        return 1
    for i, (row, sheet_row) in enumerate(zip(rows, values[1:]), start=2):
        expected = [row.get(name, "") for name in fieldnames]
        actual = ["" if v is None else str(v) for v in sheet_row]
        if expected != actual:
            print(f"export-matrix: row {i} differs from the CSV", file=sys.stderr)
            return 1
    print(f"export-matrix: spreadsheet matches the CSV ({len(rows)} rows)")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Export the traceability matrix to XLSX")
    parser.add_argument("--check", action="store_true", help="verify the spreadsheet matches the CSV")
    args = parser.parse_args()
    if args.check:
        return check()
    total, counts = build()
    print(f"export-matrix: wrote {XLSX_PATH.relative_to(ROOT)} ({total} rows)")
    for status, n in sorted(counts.items()):
        print(f"  {n:4}  {status}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
