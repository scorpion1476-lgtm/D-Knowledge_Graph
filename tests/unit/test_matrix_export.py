"""The traceability spreadsheet export.

The CSV is the source of truth, so the property that matters is that the
spreadsheet cannot drift from it. That is what these tests pin.
"""

from __future__ import annotations

import csv
import importlib.util
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed (the 'dev' extra)")

ROOT = Path(__file__).resolve().parents[2]
SCRIPT = ROOT / "scripts" / "export_matrix_xlsx.py"


@pytest.fixture(scope="module")
def exporter():
    spec = importlib.util.spec_from_file_location("export_matrix_xlsx", SCRIPT)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_committed_spreadsheet_matches_the_committed_csv(exporter):
    # The check mode returns 0 only when every cell of every row agrees.
    assert exporter.check() == 0


def test_spreadsheet_has_a_row_for_every_requirement(exporter):
    from openpyxl import load_workbook

    with exporter.CSV_PATH.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    ws = load_workbook(exporter.XLSX_PATH, read_only=True).active
    values = list(ws.values)
    assert len(values) - 1 == len(rows)
    assert list(values[0]) == list(rows[0].keys())


def test_every_status_in_the_matrix_has_a_colour(exporter):
    with exporter.CSV_PATH.open(newline="", encoding="utf-8") as fh:
        statuses = {(r["status"] or "").strip() for r in csv.DictReader(fh)}
    missing = statuses - set(exporter.STATUS_STYLE)
    assert not missing, f"no colour defined for {missing}"


def test_colour_coding_follows_the_convention(exporter):
    green = exporter.STATUS_STYLE["PRODUCTION READY"][0]
    amber = exporter.STATUS_STYLE["IMPLEMENTED BUT NOT FULLY VERIFIED"][0]
    red = exporter.STATUS_STYLE["BLOCKED BY EXTERNAL PLATFORM"][0]
    assert len({green, amber, red}) == 3, "each severity needs a distinct colour"
    assert exporter.STATUS_STYLE["NOT IMPLEMENTED"][0] == red


def test_header_is_frozen_and_filterable(exporter):
    from openpyxl import load_workbook

    ws = load_workbook(exporter.XLSX_PATH).active
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref, "the sheet must be filterable"
    assert ws.auto_filter.ref.startswith("A1:")


def test_every_column_has_a_declared_layout(exporter):
    with exporter.CSV_PATH.open(newline="", encoding="utf-8") as fh:
        fieldnames = list(csv.DictReader(fh).fieldnames or [])
    for name in fieldnames:
        assert name in exporter.COLUMN_LAYOUT, f"{name} would fall back to a generic width"


def test_export_is_reproducible_from_the_csv(exporter, tmp_path, monkeypatch):
    # Re-exporting must still agree with the CSV; a build that drifted would
    # make the committed spreadsheet unverifiable.
    target = tmp_path / "matrix.xlsx"
    monkeypatch.setattr(exporter, "XLSX_PATH", target)
    total, counts = exporter.build()
    assert target.is_file()
    assert sum(counts.values()) == total
    assert exporter.check() == 0
